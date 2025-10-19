from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def create_statistics_excel(material_id: int, material_title: str, statistics: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"
    
    headers = [
        "User ID",
        "Username",
        "Имя",
        "ФИО",
        "Компания",
        "Должность",
        "Телефон",
        "Дата и время"
    ]
    
    ws.append(headers)
    
    for stat in statistics:
        ws.append([
            stat['user_id'],
            stat['username'] or '',
            stat['first_name'] or '',
            stat['full_name'] or '',
            stat['company'] or '',
            stat['position'] or '',
            stat['phone_number'] or '',
            stat['viewed_at'].strftime('%d.%m.%Y %H:%M:%S') if stat['viewed_at'] else ''
        ])
    
    # Автоподгонка ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        ws.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_title = "".join(c for c in material_title if c.isalnum() or c in (' ', '-', '_')).rstrip()[:30]
    filename = f"stats_{material_id}_{safe_title}_{timestamp}.xlsx"
    filepath = Path(filename)
    
    wb.save(filepath)
    
    return filepath


def create_users_excel(users: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = [
        "User ID",
        "Username",
        "Имя",
        "ФИО",
        "Компания",
        "Должность",
        "Телефон",
        "Дата регистрации",
        "Просмотренные материалы"
    ]
    
    ws.append(headers)
    
    for user in users:
        materials_text = '\n'.join(user['materials']) if user['materials'] else ''
        
        ws.append([
            user['user_id'],
            user['username'] or '',
            user['first_name'] or '',
            user['full_name'] or '',
            user['company'] or '',
            user['position'] or '',
            user['phone_number'] or '',
            user['created_at'].strftime('%d.%m.%Y %H:%M:%S') if user['created_at'] else '',
            materials_text
        ])
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_lines = str(cell.value).split('\n')
                    for line in cell_lines:
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        
        if column_letter == 'I':
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
        else:
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"users_export_{timestamp}.xlsx"
    filepath = Path(filename)
    
    wb.save(filepath)
    
    return filepath

